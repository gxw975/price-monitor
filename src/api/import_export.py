"""批量导入导出 API

提供商品/关键词的 Excel 导入和列表数据导出。
权限：manager/admin 可导入，全员可导出。
"""

from __future__ import annotations

import io
import logging
import os
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse

from services.auth_service import get_current_user

load_dotenv()
logger = logging.getLogger("api.import_export")

DATABASE_URL = os.getenv("DATABASE_URL", "")
_SCHEMA = "price_monitor"

router = APIRouter(prefix="/api/import", tags=["import-export"])


def _parse_db_url(url: str) -> tuple[str, str]:
    from urllib.parse import parse_qs, urlparse, urlunparse
    parsed = urlparse(url)
    qs = parse_qs(parsed.query)
    schema = qs.get("schema", [_SCHEMA])[0]
    clean = urlunparse(parsed._replace(query=""))
    return clean, schema


def _get_conn() -> Any:
    clean, schema = _parse_db_url(DATABASE_URL)
    conn = psycopg2.connect(clean)
    with conn.cursor() as cur:
        cur.execute("SET search_path TO %s", (schema,))
    return conn


def _check_write_permission(role: str) -> None:
    if role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="权限不足，仅管理员和主管可以操作")


@router.post("/products")
def import_products(
    file: UploadFile = File(...),
    current_user: dict[str, Any] = Depends(get_current_user),
) -> dict[str, Any]:
    _check_write_permission(current_user["role"])

    if not file.filename:
        raise HTTPException(status_code=400, detail="请上传文件")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls / .csv 文件")

    try:
        content = file.file.read()
    except Exception:
        raise HTTPException(status_code=400, detail="读取文件失败")

    rows: list[dict[str, str]] = []

    if ext == "csv":
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append({k.strip(): v.strip() if v else "" for k, v in r.items()})
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d: dict[str, str] = {}
            for i, h in enumerate(headers):
                d[h] = str(row[i]).strip() if row[i] is not None and i < len(row) else ""
            if any(v for v in d.values()):
                rows.append(d)

    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有数据")

    conn = _get_conn()
    created = 0
    updated = 0
    errors: list[str] = []
    try:
        with conn:
            with conn.cursor() as cur:
                for i, r in enumerate(rows, 1):
                    pid = r.get("product_id") or r.get("商品ID") or r.get("id") or ""
                    title = r.get("title") or r.get("商品名称") or r.get("名称") or ""
                    shop = r.get("shop_name") or r.get("店铺") or r.get("店铺名称") or ""
                    shop_type = r.get("shop_type") or r.get("店铺类型") or r.get("类型") or "taobao"

                    if not pid or not title:
                        errors.append(f"第{i}行: 缺少商品ID或标题")
                        continue

                    cur.execute(
                        'SELECT product_id FROM "Product" WHERE product_id = %s',
                        (pid,),
                    )
                    if cur.fetchone():
                        cur.execute(
                            'UPDATE "Product" SET title=%s, shop_name=%s, shop_type=%s, last_updated_at=NOW() '
                            "WHERE product_id=%s",
                            (title, shop, shop_type, pid),
                        )
                        updated += 1
                    else:
                        cur.execute(
                            'INSERT INTO "Product" (product_id, title, shop_name, shop_type, is_approved) '
                            "VALUES (%s, %s, %s, %s, FALSE)",
                            (pid, title, shop, shop_type),
                        )
                        created += 1

            conn.commit()
        logger.info(
            "导入商品: created=%d updated=%d errors=%d (by %s)",
            created, updated, len(errors), current_user["username"],
        )
        return {
            "success": True,
            "created": created,
            "updated": updated,
            "errors": errors,
        }
    except Exception:
        logger.exception("导入商品失败")
        conn.rollback()
        raise HTTPException(status_code=500, detail="导入商品失败")
    finally:
        conn.close()


@router.post("/keywords")
def import_keywords(
    file: UploadFile = File(...),
    current_user: dict[str, Any] = Depends(get_current_user),
) -> dict[str, Any]:
    _check_write_permission(current_user["role"])

    if not file.filename:
        raise HTTPException(status_code=400, detail="请上传文件")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls / .csv 文件")

    try:
        content = file.file.read()
    except Exception:
        raise HTTPException(status_code=400, detail="读取文件失败")

    rows: list[dict[str, str]] = []

    if ext == "csv":
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append({k.strip(): v.strip() if v else "" for k, v in r.items()})
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d: dict[str, str] = {}
            for i, h in enumerate(headers):
                d[h] = str(row[i]).strip() if row[i] is not None and i < len(row) else ""
            if any(v for v in d.values()):
                rows.append(d)

    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有数据")

    conn = _get_conn()
    created = 0
    skipped = 0
    errors: list[str] = []
    try:
        with conn:
            with conn.cursor() as cur:
                for i, r in enumerate(rows, 1):
                    name = r.get("name") or r.get("关键词") or r.get("名称") or ""
                    platform = r.get("platform") or r.get("平台") or "taobao"

                    if not name:
                        errors.append(f"第{i}行: 缺少关键词名称")
                        continue

                    cur.execute(
                        'SELECT id FROM "Keyword" WHERE name = %s AND platform = %s',
                        (name, platform),
                    )
                    if cur.fetchone():
                        skipped += 1
                        continue

                    cur.execute(
                        'INSERT INTO "Keyword" (name, platform, created_by) '
                        "VALUES (%s, %s, %s)",
                        (name, platform, current_user["user_id"]),
                    )
                    created += 1

            conn.commit()
        logger.info(
            "导入关键词: created=%d skipped=%d errors=%d (by %s)",
            created, skipped, len(errors), current_user["username"],
        )
        return {
            "success": True,
            "created": created,
            "skipped": skipped,
            "errors": errors,
        }
    except Exception:
        logger.exception("导入关键词失败")
        conn.rollback()
        raise HTTPException(status_code=500, detail="导入关键词失败")
    finally:
        conn.close()


@router.get("/export/products")
def export_products_xlsx(
    keyword: str | None = Query(None),
    current_user: dict[str, Any] = Depends(get_current_user),
) -> StreamingResponse:
    conn = _get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if keyword:
                cur.execute(
                    'SELECT product_id, title, shop_name, shop_type, is_approved, is_whitelist, '
                    "created_at, last_updated_at FROM \"Product\" "
                    "WHERE (title ILIKE %s OR product_id ILIKE %s) "
                    "ORDER BY last_updated_at DESC LIMIT 5000",
                    (f"%{keyword}%", f"%{keyword}%"),
                )
            else:
                cur.execute(
                    'SELECT product_id, title, shop_name, shop_type, is_approved, is_whitelist, '
                    "created_at, last_updated_at FROM \"Product\" "
                    "ORDER BY last_updated_at DESC LIMIT 5000"
                )
            rows = cur.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "商品列表"
        ws.append(["商品ID", "标题", "店铺", "店铺类型", "已审核", "白名单", "创建时间", "更新时间"])

        for r in rows:
            ws.append([
                r["product_id"],
                r["title"],
                r["shop_name"],
                r.get("shop_type", ""),
                "是" if r["is_approved"] else "否",
                "是" if r["is_whitelist"] else "否",
                r["created_at"].isoformat()[:19] if r.get("created_at") else "",
                r["last_updated_at"].isoformat()[:19] if r.get("last_updated_at") else "",
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        ws.column_dimensions["B"].width = 40

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=products.xlsx"},
        )
    except Exception:
        logger.exception("导出商品失败")
        raise HTTPException(status_code=500, detail="导出商品失败")
    finally:
        conn.close()


@router.get("/export/keywords")
def export_keywords_xlsx(
    is_active: bool | None = Query(None),
    current_user: dict[str, Any] = Depends(get_current_user),
) -> StreamingResponse:
    conn = _get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            conditions: list[str] = []
            params: list[Any] = []
            if is_active is not None:
                conditions.append("k.is_active = %s")
                params.append(is_active)
            where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

            cur.execute(
                'SELECT k.name, k.platform, k.is_active, u.username AS created_by_name, '
                "k.\"createdAt\", COUNT(pk.product_id) AS product_count "
                'FROM "Keyword" k '
                'LEFT JOIN "User" u ON k.created_by = u.id '
                'LEFT JOIN "ProductKeyword" pk ON k.id = pk.keyword_id '
                f"{where} "
                "GROUP BY k.id, u.username "
                "ORDER BY k.\"createdAt\" DESC LIMIT 5000",
                params,
            )
            rows = cur.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "关键词列表"
        ws.append(["关键词", "平台", "启用", "关联商品数", "创建人", "创建时间"])

        for r in rows:
            ws.append([
                r["name"],
                r["platform"],
                "是" if r["is_active"] else "否",
                r["product_count"],
                r.get("created_by_name", ""),
                r["createdAt"].isoformat()[:10] if r.get("createdAt") else "",
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16
        ws.column_dimensions["A"].width = 30

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=keywords.xlsx"},
        )
    except Exception:
        logger.exception("导出关键词失败")
        raise HTTPException(status_code=500, detail="导出关键词失败")
    finally:
        conn.close()
